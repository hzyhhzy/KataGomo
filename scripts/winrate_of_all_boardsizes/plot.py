import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# 读取 results_nogo.txt 文件
with open("results_all.txt", "r") as file:
    lines = file.readlines()

# 解析数据
data = []
for line in lines:
    x, y, value = line.strip().split()
    value = float(value)
    processed_value = int(round((value + 100) / 2))  # 处理公式：(数字 + 100) / 2 取整
    data.append([int(x), int(y), processed_value])
    #if(x!=y):
    #    data.append([int(y), int(x), processed_value])

# 创建 DataFrame
df = pd.DataFrame(data, columns=["X", "Y", "Processed Value"])
print(df)
# 创建表格
pivot_table = df.pivot(index="Y", columns="X", values="Processed Value")

# 保存为 Excel 文件
output_file = "output_table.xlsx"
pivot_table.to_excel(output_file)

# 设置单元格颜色
def get_color(value):
    """根据值返回对应的颜色"""
    #elif value == 50:
    #    return "FFFFFF"  # 白色
    if value > 50:
        # 过渡颜色：根据值在 0 到 100 之间插值
        k = int(255 * ((value-50) / 50))
        return f"{255:02X}{255-k:02X}{255-k:02X}"
    else:
        # 过渡颜色：根据值在 0 到 100 之间插值
        k = int(255 * ((-value+50) / 50))
        return f"{255-k:02X}{255:02X}{255:02X}"

# 加载 Excel 文件并设置颜色
wb = openpyxl.load_workbook(output_file)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            color = get_color(cell.value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 保存修改后的 Excel 文件
wb.save(output_file)
print(f"表格已保存到 {output_file}")